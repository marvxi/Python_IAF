from openpyxl import load_workbook

#Festlegen der Arbeitsmappe
wb = load_workbook('xlsx datei')
for sheet in wb:
    print(sheet.title)
ws=wb.active
print('')
sheet = wb.worksheets[0]
row_count = sheet.max_row
column_count = sheet.max_column
