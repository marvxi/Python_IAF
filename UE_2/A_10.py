import openpyxl as opx

fileName = r'C:\Users\Marvi\OneDrive\03_Studium Unterlagen\10_Master\03_Semester\02_Python\02_UE/MeineDaten.xlsx'
wb = opx.load_workbook(filename=fileName)

ws = wb['Tabelle1']

# Wie lang ist das Datenset?

count = 0
field = 'A'
data = ''

while data != None:
    count+=1
    data = ws['A' + str(count)].value

# Ins dic schreiben

meineDaten = {ws['A1'].value: [],   #Dictionary anlegen mit entsprechenden key
              ws['B1'].value: [],
              ws['C1'].value: [] }

for i in (2, count-1):
    meineDaten[ws['A1'].value].append(float(ws['A' + str(i)].value))
    meineDaten[ws['B1'].value].append(float(ws['B' + str(i)].value))
    meineDaten[ws['C1'].value].append(float(ws['C' + str(i)].value))

print(meineDaten)



