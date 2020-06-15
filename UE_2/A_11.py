import openpyxl as opx
import numpy as np
# Erzeuge die x und y Werte

t = []
y = []
time = 0

while time < 25:
    time+=0.1
    t.append(time)
    y.append(np.sin(time)**2)

# Excel liste vorbereiten
wb = opx.Workbook() # Workbook objekt initialisieren
ws = wb.active  # Worksheet aktivieren
ws['A' + str(1)] = "Zeit t"
ws['B' + str(1)] = "Funktion y"
#str repräsentiert den counter für die jeweilige Zeile
for i in range(len(t)):

    ws['A' + str(i + 2)] = t[i]
    ws['B' + str(i + 2)] = y[i]

fileName = r'C:\Users\Marvi\Documents\Pycharm_Projects\Python_IAF\UE_2\A11_Sin.xlsx'
wb.save(fileName)
